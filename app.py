from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from datetime import datetime
import random
import json
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Change this to a random string

SESSION_FILE = "web_session.json"
LOG_FILE = "shift_log.txt"
EXCEL_FILE = "shifts.xlsx"
EXCEL_DOWNLOAD_PASSWORD = "EasternCC001"  # Change this to your desired password

JOB_SITES = [
    "2025 DC water - Washington DC 20032",
    "BRWRF Phase 3 Package 1 - Ashburn 20147",
    "DC Water Projects - Washington DC 20032",
    "[HMMA] Stamp Shop Roof Improvement - Montgomery, Alabama 36105",
    "Hyundai Product Center Office - Carnesville, Georgia",
    "JWA22 - Kokomo, Indiana 46901",
    "[KARR] Korean Ambassador's Residence Renovation, NW Washington 20016",
    "LGES-ALPHA Project-1F0 - Queen Creek 85140",
    "LGESMI2 EV Battery Plant Main BLDG - Holland Michigan 49423",
    "TPO Roof Project - Fairfax 22030"
]

def save_session_data(data):
    with open(SESSION_FILE, "w") as f:
        json.dump(data, f)

def load_session_data():
    if os.path.exists(SESSION_FILE):
        with open(SESSION_FILE, "r") as f:
            return json.load(f)
    return {}

def log_event(event, dt, code=None):
    with open(LOG_FILE, "a") as f:
        code_str = f" [Code: {code}]" if code else ""
        f.write(f"{dt.strftime('%Y-%m-%d %I:%M %p')} - {event}{code_str}\n")

def generate_code():
    return str(random.randint(100000, 999999))

def format_seconds(secs):
    hours = int(secs // 3600)
    minutes = int((secs % 3600) // 60)
    return f"{hours}h {minutes}m"

def write_clockin_to_excel(name, job_site, clock_in_dt, code):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Name", "Job Site", "Clock In", "Clock Out", "Total Time", "Working Time", "Breaks", "Code"
        ])
        wb.save(EXCEL_FILE)
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        name,
        job_site,
        clock_in_dt.strftime('%Y-%m-%d %I:%M %p'),
        "", "", "", "",  # clock out, total, working, breaks
        code
    ])
    wb.save(EXCEL_FILE)

def update_clockout_in_excel(code, clock_out_dt, total_time, working_time, breaks_str):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):  # skip header
        if str(row[7].value) == str(code) and not row[3].value:
            row[3].value = clock_out_dt.strftime('%Y-%m-%d %I:%M %p')
            row[4].value = total_time
            row[5].value = working_time
            row[6].value = breaks_str
            break
    wb.save(EXCEL_FILE)

@app.route("/", methods=["GET", "POST"])
def index():
    data = load_session_data()
    clocked_in = data.get("clocked_in", False)
    on_break = data.get("on_break", False)
    code = data.get("code", None)
    clock_in_time = data.get("clock_in_time")
    breaks = data.get("breaks", [])
    name = data.get("name", "")
    job_site = data.get("job_site", "")
    status = ""
    show_code = False

    if request.method == "POST":
        action = request.form.get("action")
        now = datetime.now()

        if action == "clockin" and not clocked_in:
            name = request.form.get("name", "").strip()
            job_site = request.form.get("job_site", "")
            if not name:
                flash("Please enter your name.", "error")
                return redirect(url_for("index"))
            if not job_site:
                flash("Please select a job site.", "error")
                return redirect(url_for("index"))
            code = generate_code()
            log_event("Clocked In", now, code)
            clocked_in = True
            on_break = False
            clock_in_time = now.isoformat()
            breaks = []
            write_clockin_to_excel(name, job_site, now, code)
            flash(
                f"Your code is: <b>{code}</b><br>"
                f"<span style='color:red;'>This code is required to clock out. Please write it down or remember it. It will not be shown again!</span>",
                "success"
            )
            status = "Clocked in."
            show_code = False  # Do not show code in template
        elif action == "break" and clocked_in and not on_break:
            log_event("Break Start", now, code)
            on_break = True
            breaks.append({"start": now.isoformat(), "end": None})
            status = "Break started."
            show_code = True
        elif action == "resume" and clocked_in and on_break:
            log_event("Break End", now, code)
            on_break = False
            if breaks and breaks[-1]["end"] is None:
                breaks[-1]["end"] = now.isoformat()
            status = "Break ended."
            show_code = True
        elif action == "clockout" and clocked_in:
            input_code = request.form.get("input_code")
            if input_code != code:
                flash("Incorrect code.", "error")
                return redirect(url_for("index"))
            if on_break:
                log_event("Break End (auto on clock out)", now, code)
                on_break = False
                if breaks and breaks[-1]["end"] is None:
                    breaks[-1]["end"] = now.isoformat()
            log_event("Clocked Out", now, code)
            clocked_in = False

            # Calculate shift times
            if clock_in_time:
                clock_in_dt = datetime.fromisoformat(clock_in_time)
                clock_out_dt = now
                total_time = clock_out_dt - clock_in_dt
                total_break = sum(
                    (datetime.fromisoformat(b["end"]) - datetime.fromisoformat(b["start"]))
                    .total_seconds() for b in breaks if b["end"] is not None
                )
                working_time = total_time.total_seconds() - total_break
                breaks_str = "; ".join(
                    f"{datetime.fromisoformat(b['start']).strftime('%I:%M %p')} - {datetime.fromisoformat(b['end']).strftime('%I:%M %p')}"
                    for b in breaks if b["end"] is not None
                )
                update_clockout_in_excel(
                    code,
                    clock_out_dt,
                    format_seconds(total_time.total_seconds()),
                    format_seconds(working_time),
                    breaks_str
                )
                flash(
                    f"Shift complete!<br>"
                    f"Total time: <b>{format_seconds(total_time.total_seconds())}</b><br>"
                    f"Actual working time: <b>{format_seconds(working_time)}</b>",
                    "success"
                )
            else:
                flash("Shift complete!", "success")

            code = None
            clock_in_time = None
            breaks = []
            name = ""
            job_site = ""
            show_code = False
        else:
            flash("Invalid action or state.", "error")
            return redirect(url_for("index"))

        save_session_data({
            "clocked_in": clocked_in,
            "on_break": on_break,
            "code": code,
            "clock_in_time": clock_in_time,
            "breaks": breaks,
            "name": name,
            "job_site": job_site
        })
        if action != "clockout":
            flash(status, "success")
        return redirect(url_for("index"))

    return render_template("index.html", 
                         clocked_in=clocked_in, 
                         on_break=on_break, 
                         code=code, 
                         job_sites=JOB_SITES,
                         name=name,
                         job_site=job_site)

@app.route("/download-excel", methods=["GET", "POST"])
def download_excel():
    if request.method == "POST":
        password = request.form.get("password", "")
        if password == EXCEL_DOWNLOAD_PASSWORD:
            if not os.path.exists(EXCEL_FILE):
                flash("No Excel file found.", "error")
                return redirect(url_for("index"))
            return send_file(EXCEL_FILE, as_attachment=True)
        else:
            flash("Incorrect password.", "error")
            return render_template("download_excel.html")
    return render_template("download_excel.html")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port, debug=True)
